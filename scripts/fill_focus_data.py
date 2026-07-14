#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Скрипт 2: Заполнение фокусов данными из таблицы для Hearts of Iron IV / Script 2: Fill focuses with data from spreadsheet for Hearts of Iron IV
"""

# ============================================================
# НАСТРОЙКИ / SETTINGS - ИЗМЕНЯЙТЕ ЗНАЧЕНИЯ / EDIT VALUES HERE
# ============================================================

# Путь к файлу дерева фокусов (.txt) / Path to focus tree file (.txt)
FOCUS_FILE_PATH = r"C:/Users/Admin/Documents/Paradox Interactive/Hearts of Iron IV/mod/sonic017/common/national_focus/rcr_eggman.txt"

# Путь к файлу таблицы (.xlsx, .xls, .csv) / Path to spreadsheet file (.xlsx, .xls, .csv)
TABLE_FILE_PATH = r"C:/Users/Admin/Desktop/Great_Patriotic_War_Focus_Tree.xlsx"

# Имя листа в таблице (если не указан - берётся активный лист)
SHEET_NAME = "RCR"

# Столбец в таблице, содержащий id фокусов (для сопоставления)
ID_COLUMN = "Technical focus name"

# Словарь сопоставления: "поле_в_файле" : "столбец_в_таблице"
MAPPING_DICT = {
    "completion_reward": "Effects",
    # "icon": "Icon Status",
    # "cost": "Cost",
}

# ============================================================
# КОД СКРИПТА - НИЧЕГО НЕ МЕНЯТЬ / SCRIPT CODE - DO NOT EDIT
# ============================================================

import os
import re
import sys

def remove_commas(text):
    """Удаляет все запятые из текста / Removes all commas from text"""
    if text is None:
        return None
    if not isinstance(text, str):
        text = str(text)
    return text.replace(',', '')

def read_spreadsheet(file_path, sheet_name):
    """Читает таблицу и возвращает словарь {id: {column: value}}"""
    try:
        if file_path.endswith('.csv'):
            import csv
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        else:
            try:
                import openpyxl
            except ImportError:
                print("[ERROR] Для .xlsx/.xls файлов нужна библиотека openpyxl")
                print("[ERROR] Установите: pip install openpyxl")
                sys.exit(1)
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"[INFO] Используется лист / Using sheet: {sheet_name}")
            else:
                ws = wb.active
                print(f"[INFO] Используется активный лист / Using active sheet: {ws.title}")
            
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val) if val else f"Column_{col}")
            
            rows = []
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    val = ws.cell(row=row, column=col_idx).value
                    row_data[header] = str(val) if val is not None else ""
                if any(v.strip() for v in row_data.values()):
                    rows.append(row_data)
        
        result = {}
        for row in rows:
            focus_id = row.get(ID_COLUMN, "").strip()
            if focus_id:
                result[focus_id] = row
        return result
    
    except FileNotFoundError:
        print(f"[ERROR] Файл не найден / File not found: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] Ошибка чтения таблицы / Error reading spreadsheet: {e}")
        sys.exit(1)

def read_focus_file(file_path):
    """Читает файл фокусов"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        print(f"[ERROR] Файл не найден / File not found: {file_path}")
        sys.exit(1)

def extract_focus_blocks_robust(content):
    """
    Механически находит границы блоков 'focus = { ... }' по балансу фигурных скобок.
    Это исключает захват чужих закрывающих скобок дерева фокусов.
    """
    focuses = []
    # Ищем все вхождения 'focus = {' с любыми пробелами/отступами
    pattern = r'focus\s*=\s*\{'
    
    for match in re.finditer(pattern, content):
        start_idx = match.start()
        # Ищем внутренности скобок, считая баланс
        brace_level = 0
        end_idx = -1
        
        for idx in range(start_idx, len(content)):
            char = content[idx]
            if char == '{':
                brace_level += 1
            elif char == '}':
                brace_level -= 1
                if brace_level == 0:
                    end_idx = idx + 1
                    break
        
        if end_idx != -1:
            full_block = content[start_idx:end_idx]
            # Вытаскиваем ID фокуса из этого конкретного блока
            id_match = re.search(r'^\s*id\s*=\s*(\S+)', full_block, re.MULTILINE)
            if id_match:
                focus_id = id_match.group(1)
                focuses.append({
                    'id': focus_id,
                    'full_block': full_block,
                    'start_idx': start_idx,
                    'end_idx': end_idx
                })
    return focuses

def inject_field_value(focus_block, field_name, raw_value):
    """
    Механически находит field_name = { ... } внутри блока фокуса 
    и заменяет его содержимое, сохраняя внешнюю структуру.
    """
    if raw_value is None or str(raw_value).strip() == "":
        clean_value = "# TODO"
    else:
        clean_value = remove_commas(str(raw_value))

    # Регулярка ищет field_name = { с любыми отступами
    pattern = rf'(\s*{field_name}\s*=\s*\{{)'
    match = re.search(pattern, focus_block)
    
    if not match:
        # Если поля нет — скрипт НИЧЕГО не делает (как вы и просили)
        return focus_block
    
    # Индекс начала внутренностей скобок после '{'
    content_start = match.end()
    
    # Считаем баланс скобок строго для этого поля, чтобы найти его закрывающую скобку
    brace_level = 1
    content_end = -1
    for idx in range(content_start, len(focus_block)):
        char = focus_block[idx]
        if char == '{':
            brace_level += 1
        elif char == '}':
            brace_level -= 1
            if brace_level == 0:
                content_end = idx
                break
                
    if content_end != -1:
        # Собираем блок заново: всё до открытия + значение из таблицы + всё после закрытия
        # Сохраняем красивые переносы строк
        before = focus_block[:content_start]
        after = focus_block[content_end:]
        
        # Определяем базовый отступ для текста внутри
        indent_match = re.search(r'^\s*', match.group(0))
        base_indent = indent_match.group(0) if indent_match else "\t\t\t"
        if not base_indent.startswith('\n'):
            base_indent = "\n" + base_indent + "\t"
            
        return f"{before}{base_indent}{clean_value}\n{after}"
        
    return focus_block

def write_focus_file(file_path, content):
    """Записывает обновлённое содержимое в файл"""
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"[INFO] Файл сохранён / File saved: {file_path}")

def main():
    print("=" * 60)
    print("[INFO] Механическое заполнение полей фокусов")
    print("=" * 60)
    
    # 1. Читаем таблицу
    table_data = read_spreadsheet(TABLE_FILE_PATH, SHEET_NAME)
    print(f"[INFO] Найдено записей в таблице: {len(table_data)}")
    if not table_data:
        print("[ERROR] Таблица пуста.")
        return
        
    # 2. Читаем файл фокусов
    file_content = read_focus_file(FOCUS_FILE_PATH)
    
    # 3. Извлекаем блоки фокусов по точному балансу скобок
    focuses = extract_focus_blocks_robust(file_content)
    print(f"[INFO] Найдено фокусов в файле: {len(focuses)}")
    
    if not focuses:
        print("[WARN] Скрипт не смог найти блоки 'focus = { ... }'. Проверьте файл.")
        return

    # 4. Двигаемся по файлу СЗАДИ НАПЕРЕД, чтобы при изменении длины текста 
    # не съезжали индексы (start_idx / end_idx) еще не измененных блоков
    working_content = file_content
    matched_count = 0
    skipped_count = 0
    
    for focus in reversed(focuses):
        focus_id = focus['id']
        
        if focus_id not in table_data:
            skipped_count += 1
            continue
            
        matched_count += 1
        row_data = table_data[focus_id]
        
        # Берем чистый оригинальный блок фокуса из файла
        current_block = focus['full_block']
        
        # Механически заменяем только то, что указано в MAPPING_DICT
        for file_field, table_column in MAPPING_DICT.items():
            raw_value = row_data.get(table_column, "")
            current_block = inject_field_value(current_block, file_field, raw_value)
            
        # Вклеиваем измененный блок обратно на свое место в файле
        working_content = working_content[:focus['start_idx']] + current_block + working_content[focus['end_idx']:]

    print(f"[INFO] Обработано: {matched_count + skipped_count} | Совпало: {matched_count} | Пропущено: {skipped_count}")
    
    if matched_count > 0:
        # Сохраняем файл "как есть", без финальных сортировок строк
        write_focus_file(FOCUS_FILE_PATH, working_content)
        print("[SUCCESS] Операция успешно завершена.")
    else:
        print("[WARN] Ни одного совпадения не найдено. Файл не изменен.")

if __name__ == "__main__":
    main()