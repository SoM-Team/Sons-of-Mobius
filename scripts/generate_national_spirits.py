#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Скрипт 4: Создание национальных духов из таблицы для Hearts of Iron IV / Script 4: Generate national spirits from spreadsheet for Hearts of Iron IV

Что делает скрипт / What this script does:
- Читает таблицу и создаёт национальные духи в указанном файле / Reads spreadsheet and creates national spirits in the specified file
- Каждый дух имеет имя (префикс + название), иконку, модификаторы и опциональные блоки / Each spirit has name (prefix + name), icon, modifiers and optional blocks

Как использовать / How to use:
1. Сделать копию скрипта и не использовать оригинальный скрипт / Make a copy of this script and do not use the original script
2. Заполнить таблицу данными / Fill spreadsheet with data
3. Настроить параметры в блоке "НАСТРОЙКИ" ниже / Define parameters in "SETTINGS" block below
4. Запустить скрипт: python generate_national_spirits.py / Launch the script: python generate_national_spirits.py
"""

# ============================================================
# НАСТРОЙКИ / SETTINGS - ИЗМЕНЯЙТЕ ЗНАЧЕНИЯ / EDIT VALUES HERE
# ============================================================

# Путь к файлу национальных духов (.txt) / Path to national spirits file (.txt)
SPIRIT_FILE_PATH = r"C:/Users/Admin/Documents/Paradox Interactive/Hearts of Iron IV/mod/sonic017/common/ideas/UWM.txt"

# Путь к файлу таблицы (.xlsx, .xls, .csv) / Path to spreadsheet file
TABLE_FILE_PATH = r"C:/Users/Admin/Desktop/United_Workers_of_Mobius_Focus_Tree.xlsx"

# Имя листа в таблице (если не указан - берётся активный лист)
# Sheet name in spreadsheet (if empty - takes active sheet)
SHEET_NAME = "UWM"  # <--- ИЗМЕНИТЕ ПРИ НЕОБХОДИМОСТИ

# Столбец в таблице с названием духа / Column with spirit name
NAME_COLUMN = "National Ideas Names"

# Столбец в таблице с эффектами (модификаторами) / Column with effects (modifiers)
EFFECTS_COLUMN = "National Ideas Effects"

# Префикс для названия духа / Prefix for spirit name
# Пример: "NOR_" -> NOR_Amadeus_The_First_Democratic_Reform
PREFIX = "UWM_"

# Изображение по умолчанию / Default picture
# Пример: GFX_idea_generic
DEFAULT_PICTURE = "GFX_idea_generic"

# available блок (если не пустой - вставляется как есть)
# available block (if not empty - inserted as-is)
# Если оставить пустой - блок не создаётся / If left empty - block not created

AVAILABLE_TEXT = ""

# Пример / Example: 
# """
# 				tag = NOR
# 				OR = {
# 					has_war_with = KOA
# 					has_war_with = UWM
# 					has_war_with = SOT
# 				}"""

# allowed блок (если не пустой - вставляется как есть)
# allowed block (if not empty - inserted as-is)
# Если оставить пустой - блок не создаётся / If left empty - block not created

ALLOWED_TEXT = ""

# Пример / Example: 
# """
# 				tag = NOR"""

# cancel_if_invalid - дух удаляется при невыполнении условий / spirit removed when conditions not met
# True / False

CANCEL_IF_INVALID = False

# ============================================================
# КОД СКРИПТА - НИЖЕ НИЧЕГО НЕ МЕНЯТЬ / SCRIPT CODE - DO NOT EDIT BELOW
# ============================================================

import os
import re
import sys

def ensure_directory_exists(file_path):
    """Проверяет/создаёт директорию для файла"""
    directory = os.path.dirname(file_path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory)
        print(f"[INFO] Создана директория / Created directory: {directory}")

def read_spreadsheet(file_path, sheet_name):
    """Читает таблицу и возвращает список строк"""
    try:
        if file_path.endswith('.csv'):
            import csv
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        else:
            try:
                import openpyxl
            except ImportError:
                print("[ERROR] Для .xlsx/.xls файлов нужна библиотека openpyxl")
                print("[ERROR] Установите: pip install openpyxl")
                print("[ERROR] Или используйте .csv формат")
                sys.exit(1)
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Выбираем лист
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"[INFO] Используется лист / Using sheet: {sheet_name}")
            else:
                ws = wb.active
                if sheet_name:
                    print(f"[WARN] Лист '{sheet_name}' не найден, используется активный: {ws.title}")
                else:
                    print(f"[INFO] Используется активный лист / Using active sheet: {ws.title}")
            
            # Читаем заголовки
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers.append(str(val))
                else:
                    headers.append(f"Column_{col}")
            
            # Читаем данные
            rows = []
            for row in range(2, ws.max_row + 1):
                row_data = {}
                has_data = False
                for col_idx, header in enumerate(headers, start=1):
                    val = ws.cell(row=row, column=col_idx).value
                    if val is not None and str(val).strip():
                        row_data[header] = str(val)
                        has_data = True
                    else:
                        row_data[header] = ""
                if has_data:
                    rows.append(row_data)
        
        return rows
    
    except FileNotFoundError:
        print(f"[ERROR] Файл не найден / File not found: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] Ошибка чтения таблицы / Error reading spreadsheet: {e}")
        sys.exit(1)

def read_existing_spirits(file_path):
    """
    Читает существующий файл и извлекает имена уже существующих духов
    Returns: set of spirit ids, full file content
    """
    if not os.path.exists(file_path):
        return set(), ""
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Ищем все id духов в файле (после "country = {" и до "}")
    # Паттерн для поиска: строки вида \t\tidea_name = {
    pattern = r'^\t\t([A-Za-z0-9_]+)\s*=\s*\{'
    matches = re.findall(pattern, content, re.MULTILINE)
    spirit_ids = set(matches)
    
    return spirit_ids, content

def format_modifier(effects_text):
    """
    Форматирует текст эффектов в блок modifier
    """
    if not effects_text or effects_text.strip() == "":
        return "\t\t\tmodifier = { }"
    
    # Убираем лишние запятые
    content = effects_text.replace(',', '')
    
    # Разбиваем на строки и добавляем отступы
    lines = content.strip().split('\n')
    formatted_lines = []
    for line in lines:
        if line.strip():
            formatted_lines.append(f"\t\t\t\t{line.strip()}")
    
    if formatted_lines:
        return "\t\t\tmodifier = {\n" + "\n".join(formatted_lines) + "\n\t\t\t}"
    else:
        return "\t\t\tmodifier = { }"

def generate_spirit_block(spirit_id, picture, modifier_text, available_text, allowed_text, cancel_if_invalid):
    """
    Генерирует полный блок национального духа
    Структура:
        idea_name = {
            allowed = { ... }
            available = { ... }
            picture = ...
            cancel_if_invalid = yes/no
            modifier = { ... }
        }
    """
    lines = []
    lines.append(f"\t\t{spirit_id} = {{")
    
    # allowed блок (если задан)
    if allowed_text and allowed_text.strip():
        lines.append("\t\t\tallowed = {")
        for line in allowed_text.strip().split('\n'):
            if line.strip():
                lines.append(f"\t\t\t\t{line.strip()}")
        lines.append("\t\t\t}")
    
    # available блок (если задан)
    if available_text and available_text.strip():
        lines.append("\t\t\tavailable = {")
        for line in available_text.strip().split('\n'):
            if line.strip():
                lines.append(f"\t\t\t\t{line.strip()}")
        lines.append("\t\t\t}")
    
    # picture
    lines.append(f"\t\t\tpicture = {picture}")
    
    # cancel_if_invalid
    if cancel_if_invalid:
        lines.append("\t\t\tcancel_if_invalid = yes")
    else:
        lines.append("\t\t\tcancel_if_invalid = no")
    
    # modifier
    lines.append(modifier_text)
    
    lines.append("\t\t}")
    return '\n'.join(lines)

def add_spirits_to_file(existing_content, new_spirits_blocks):
    """
    Добавляет новые духи в существующий файл
    Правильная структура: ideas = { country = { ... } }
    """
    if not existing_content:
        # Файл пустой или не существует - создаём с нуля
        lines = []
        lines.append("ideas = {")
        lines.append("\tcountry = {")
        lines.append("")
        for block in new_spirits_blocks:
            lines.append(block)
            lines.append("")
        lines.append("\t}")
        lines.append("}")
        return '\n'.join(lines)
    
    # Файл существует - ищем структуру
    lines = existing_content.split('\n')
    
    # Ищем строку с открывающей скобкой для "country = {"
    country_index = -1
    for i, line in enumerate(lines):
        if re.match(r'^\tcountry\s*=\s*\{', line):
            country_index = i
            break
    
    if country_index == -1:
        # Нет блока country - нужно создать
        # Ищем последнюю закрывающую скобку (должна быть от ideas)
        # Находим последнюю строку, которая содержит "}" и не является комментарием
        closing_index = -1
        for i in range(len(lines) - 1, -1, -1):
            if lines[i].strip() == "}":
                closing_index = i
                break
        
        if closing_index == -1:
            # Странный файл - просто добавляем в конец
            return existing_content + "\n\n\tcountry = {\n" + "\n".join(new_spirits_blocks) + "\n\t}\n}"
        
        # Вставляем новый блок перед последней закрывающей скобкой
        new_lines = lines[:closing_index]
        new_lines.append("\tcountry = {")
        new_lines.append("")
        for block in new_spirits_blocks:
            new_lines.append(block)
            new_lines.append("")
        new_lines.append("\t}")
        # Сохраняем последнюю закрывающую скобку
        new_lines.append(lines[closing_index])
        return '\n'.join(new_lines)
    
    # Нашли блок country - нужно найти его закрывающую скобку
    brace_level = 0
    closing_index = -1
    for i in range(country_index, len(lines)):
        line = lines[i]
        brace_level += line.count('{')
        brace_level -= line.count('}')
        if brace_level == 0 and i > country_index:
            closing_index = i
            break
    
    if closing_index == -1:
        # Не нашли закрывающую скобку - просто добавляем в конец файла
        return existing_content + "\n\n" + "\n".join(new_spirits_blocks) + "\n"
    
    # Вставляем новые блоки перед закрывающей скобкой
    new_lines = lines[:closing_index]
    new_lines.append("")
    for block in new_spirits_blocks:
        new_lines.append(block)
        new_lines.append("")
    # Сохраняем закрывающую скобку
    new_lines.append(lines[closing_index])
    # Добавляем остальные строки после закрывающей скобки (если есть)
    if closing_index + 1 < len(lines):
        new_lines.extend(lines[closing_index + 1:])
    
    return '\n'.join(new_lines)

def clean_spirit_name(raw_name):
    """Очищает имя духа от недопустимых символов"""
    clean = raw_name.replace(' ', '_')
    clean = clean.replace('.', '')
    clean = clean.replace(',', '')
    clean = clean.replace('-', '_')
    clean = clean.replace("'", "")
    clean = clean.replace('"', '')
    clean = clean.replace('(', '')
    clean = clean.replace(')', '')
    return clean

def main():
    """Основная функция"""
    print("=" * 60)
    print("Скрипт 4: Создание национальных духов из таблицы")
    print("Script 4: Generate national spirits from spreadsheet")
    print("=" * 60)
    print()
    
    # Проверка настроек
    print("[INFO] Настройки / Settings:")
    print(f"       - Файл духов / Spirit file: {SPIRIT_FILE_PATH}")
    print(f"       - Файл таблицы / Table file: {TABLE_FILE_PATH}")
    print(f"       - Лист таблицы / Sheet name: {SHEET_NAME if SHEET_NAME else '(активный / active)'}")
    print(f"       - Столбец названий / Name column: {NAME_COLUMN}")
    print(f"       - Столбец эффектов / Effects column: {EFFECTS_COLUMN}")
    print(f"       - Префикс / Prefix: {PREFIX}")
    print(f"       - Картинка / Picture: {DEFAULT_PICTURE}")
    print(f"       - cancel_if_invalid: {CANCEL_IF_INVALID}")
    print()
    
    # Читаем существующие духи
    print("[INFO] Чтение существующего файла духов / Reading existing spirits file...")
    existing_ids, existing_content = read_existing_spirits(SPIRIT_FILE_PATH)
    print(f"[INFO] Существующих духов в файле: {len(existing_ids)}")
    print()
    
    # Читаем таблицу
    print("[INFO] Чтение таблицы / Reading spreadsheet...")
    rows = read_spreadsheet(TABLE_FILE_PATH, SHEET_NAME)
    print(f"[INFO] Найдено строк в таблице: {len(rows)}")
    print()
    
    # Извлекаем названия и эффекты
    new_spirits = []
    skipped_duplicates = 0
    
    for i, row in enumerate(rows):
        name = row.get(NAME_COLUMN, "").strip()
        effects = row.get(EFFECTS_COLUMN, "").strip()
        
        if not name:
            print(f"[WARN] Строка {i+2}: нет названия, пропущена / no name, skipped")
            continue
        
        if not effects:
            print(f"[WARN] Строка {i+2}: '{name}' - нет эффектов, пропущена / no effects, skipped")
            continue
        
        # Генерируем ID духа
        clean_name = clean_spirit_name(name)
        spirit_id = PREFIX + clean_name
        
        # Проверяем на дубликат
        if spirit_id in existing_ids:
            print(f"[SKIP] Строка {i+2}: '{name}' -> {spirit_id} уже существует / already exists")
            skipped_duplicates += 1
            continue
        
        new_spirits.append({
            'id': spirit_id,
            'name': name,
            'effects': effects
        })
        print(f"[OK] Строка {i+2}: {name} -> {spirit_id}")
    
    print()
    print(f"[INFO] Новых духов для добавления: {len(new_spirits)}")
    print(f"[INFO] Пропущено дубликатов: {skipped_duplicates}")
    print()
    
    if len(new_spirits) == 0:
        print("[INFO] Нет новых духов для добавления / No new spirits to add")
        return
    
    # Генерируем блоки для новых духов
    print("[INFO] Генерация блоков духов / Generating spirit blocks...")
    new_blocks = []
    for spirit in new_spirits:
        modifier_text = format_modifier(spirit['effects'])
        block = generate_spirit_block(spirit['id'], DEFAULT_PICTURE, modifier_text,
                                       AVAILABLE_TEXT, ALLOWED_TEXT, CANCEL_IF_INVALID)
        new_blocks.append(block)
    
    # Добавляем в существующий файл
    print("[INFO] Добавление духов в файл / Adding spirits to file...")
    new_content = add_spirits_to_file(existing_content, new_blocks)
    
    # Создаём директорию при необходимости
    ensure_directory_exists(SPIRIT_FILE_PATH)
    
    # Записываем файл
    write_spirit_file(SPIRIT_FILE_PATH, new_content)
    
    print()
    print(f"[SUCCESS] Добавлено духов / Spirits added: {len(new_spirits)}")
    print("=" * 60)

def write_spirit_file(file_path, content):
    """Записывает содержимое в файл"""
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"[INFO] Файл сохранён / File saved: {file_path}")

if __name__ == "__main__":
    main()